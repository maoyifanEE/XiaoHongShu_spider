from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from .db import Database
from .time_utils import now_shanghai
from .utils import safe_filename


PROFILE_HEADERS = ["字段", "当前值", "原始显示", "是否精确", "采集时间", "备注"]
NOTE_HEADERS = [
    "序号", "note_id", "笔记URL", "是否置顶", "标题", "帖子类型", "发布时间", "最后更新时间",
    "点赞数", "点赞原始显示", "点赞是否精确", "收藏数", "收藏原始显示", "收藏是否精确",
    "评论数", "评论原始显示", "评论是否精确",
    "正文", "标签",
    "笔记采集时间", "采集状态", "备注",
]


def export_excel(db: Database, base_dir: Path, creator_id: str, creator_name: str, logger: logging.Logger) -> Path:
    timestamp = now_shanghai().strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_filename(creator_name)}_小红书公开信息_{timestamp}.xlsx"
    output_dir = base_dir / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    final_path = output_dir / filename
    tmp_path = output_dir / f".{filename}.tmp.xlsx"

    wb = Workbook()
    ws_profile = wb.active
    ws_profile.title = "博主主页"
    ws_notes = wb.create_sheet("公开笔记")

    _write_profile_sheet(ws_profile, db.latest_profile(creator_id))
    _write_notes_sheet(ws_notes, db, creator_id)
    _style_sheet(ws_profile)
    _style_sheet(ws_notes)

    wb.save(tmp_path)
    _verify_excel(tmp_path, creator_id, len(db.current_notes(creator_id)))
    tmp_path.replace(final_path)
    logger.info("EXCEL exported path=%s", final_path)
    return final_path


def _write_profile_sheet(ws: Any, profile: Any) -> None:
    ws.append(PROFILE_HEADERS)
    if not profile:
        ws.append(["采集状态", "未找到主页快照", None, None, None, "可先运行采集模式"])
        return
    rows = [
        ("采集时间", profile["captured_at"], None, True, profile["captured_at"], None),
        ("昵称", profile["nickname"], None, True, profile["captured_at"], None),
        ("小红书号", profile["xhs_id"], None, True if profile["xhs_id"] else None, profile["captured_at"], None),
        ("user_id", profile["creator_id"], None, True, profile["captured_at"], None),
        ("canonical 主页 URL", profile["canonical_url"], None, True, profile["captured_at"], None),
        ("头像 URL", profile["avatar_url"], None, True if profile["avatar_url"] else None, profile["captured_at"], None),
        ("IP 属地", profile["ip_location"], None, True if profile["ip_location"] else None, profile["captured_at"], None),
        ("完整简介", profile["bio"], None, True if profile["bio"] else None, profile["captured_at"], None),
        ("主页标签", _json_text(profile["profile_tags"]), None, True, profile["captured_at"], None),
        ("身份标签", _json_text(profile["identity_tags"]), None, True, profile["captured_at"], None),
        ("关注数", profile["following_value"], profile["following_raw"], _bool_text(profile["following_is_exact"]), profile["captured_at"], None),
        ("粉丝数", profile["followers_value"], profile["followers_raw"], _bool_text(profile["followers_is_exact"]), profile["captured_at"], None),
        ("获赞与收藏", profile["total_interactions_value"], profile["total_interactions_raw"], _bool_text(profile["total_interactions_is_exact"]), profile["captured_at"], None),
        ("页面明确公开的性别", profile["gender"], None, True if profile["gender"] else None, profile["captured_at"], "未推测"),
    ]
    for row in rows:
        ws.append(list(row))


def _write_notes_sheet(ws: Any, db: Database, creator_id: str) -> None:
    ws.append(NOTE_HEADERS)
    for index, note in enumerate(db.current_notes(creator_id), start=1):
        ws.append(
            [
                index,
                note["note_id"],
                note["canonical_url"],
                _bool_text(note["is_pinned"]),
                note["title"],
                note["note_type"],
                note["publish_time"],
                note["updated_time"],
                note["likes_value"],
                note["likes_raw"],
                _bool_text(note["likes_is_exact"]),
                note["collects_value"],
                note["collects_raw"],
                _bool_text(note["collects_is_exact"]),
                note["comments_value"],
                note["comments_raw"],
                _bool_text(note["comments_is_exact"]),
                note["body"],
                _json_text(note["hashtags"]),
                note["metrics_captured_at"],
                note["status"],
                note["status_note"],
            ]
        )


def _style_sheet(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column_letter] = min(max(widths.get(cell.column_letter, 0), len(value) + 2), 60)
            if len(value) > 30 or "\n" in value:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, width in widths.items():
        ws.column_dimensions[col].width = max(width, 10)
    if ws.max_row >= 2 and ws.max_column >= 1:
        table = Table(displayName=f"Table_{ws.title}", ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)


def _verify_excel(path: Path, creator_id: str, expected_notes: int) -> None:
    wb = load_workbook(path)
    assert "博主主页" in wb.sheetnames
    assert "公开笔记" in wb.sheetnames
    ws = wb["公开笔记"]
    headers = [cell.value for cell in ws[1]]
    missing = [header for header in NOTE_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Excel 缺少关键列: {missing}")
    note_rows = max(ws.max_row - 1, 0)
    if note_rows != expected_notes:
        raise ValueError(f"Excel 笔记行数不一致: expected={expected_notes} actual={note_rows} creator_id={creator_id}")
    ids = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)]
    if len(ids) != len(set(ids)):
        raise ValueError("Excel note_id 存在重复")


def _json_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        try:
            loaded = json.loads(value)
            if isinstance(loaded, list):
                return " ".join(str(item) for item in loaded)
        except json.JSONDecodeError:
            return value
        return value
    if isinstance(value, list):
        return " ".join(str(item) for item in value)
    return str(value)


def _bool_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return "是" if bool(value) else "否"
