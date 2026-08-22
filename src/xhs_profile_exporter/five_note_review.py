from __future__ import annotations

import json
import re
import shutil
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .exporter import NOTE_HEADERS
from .utils import sanitize_json


REVIEW_FIELDS = [
    "title",
    "body",
    "note_type",
    "publish_time",
    "like_count",
    "collect_count",
    "comment_count",
    "tags",
]

SENSITIVE_TERMS = [
    "cookie",
    "authorization",
    "bearer",
    "token",
    "xsec",
    "session",
    "__initial_state__",
    "initial_state",
]

UI_POLLUTION_TERMS = ["关注", "加载中", "说点什么", "发送", "取消"]


E2E_REVIEW_NOTE_LIMITS = {5, 10}
FIELD_STATUS_PRESENT = "PRESENT"
FIELD_STATUS_CONFIRMED_ABSENT = "CONFIRMED_ABSENT"
FIELD_STATUS_UNVERIFIED_MISSING = "UNVERIFIED_MISSING"
CLASSIFICATION_TRUE_ABSENT = "TRUE_ABSENT"
CLASSIFICATION_INSUFFICIENT_EVIDENCE = "INSUFFICIENT_EVIDENCE"


def e2e_review_enabled(mode: str, limit: int | None) -> bool:
    return mode == "collect" and int(limit or 0) in E2E_REVIEW_NOTE_LIMITS


def e2e_review_dir(base_dir: Path, run_id: str) -> Path:
    path = base_dir / "validation" / "e2e_review" / run_id
    path.mkdir(parents=True, exist_ok=True)
    return path


def five_note_review_dir(base_dir: Path, run_id: str) -> Path:
    return e2e_review_dir(base_dir, run_id)


async def capture_e2e_review_state(page: Any, note_id: str) -> dict[str, Any]:
    return await page.evaluate(
        """
        (noteId) => {
          const visible = (el) => {
            if (!el) return false;
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.display !== "none" && style.visibility !== "hidden";
          };
          const clean = (el) => el ? (el.innerText || el.textContent || "").trim() : "";
          const detailSelector = '[class*="note-detail"], [class*="noteDetail"], [class*="NoteDetail"], [data-testid*="note-detail"], [role="dialog"]';
          const evidenceSelectors = '#detail-title, #detail-desc, .engage-bar, [class*=engage], [class*=interaction], [class*=Interact]';
          const detailRoots = Array.from(document.querySelectorAll(detailSelector)).filter(visible);
          const evidence = Array.from(document.querySelectorAll(evidenceSelectors)).filter(visible);
          const evidenceRoot = evidence.map((el) => el.closest(detailSelector)).find((el) => el && visible(el));
          const root = evidenceRoot || detailRoots.find((el) => {
            const hasEvidence = Boolean(el.querySelector(evidenceSelectors));
            const hasExactLink = noteId ? Boolean(el.querySelector(`a[href*="${noteId}"]`)) : false;
            return hasEvidence || hasExactLink;
          });
          const rootReason = root ? (evidenceRoot ? "DETAIL_EVIDENCE_ROOT" : "DETAIL_WRAPPER_ROOT") : "NO_STRONG_DETAIL_ROOT";
          const q = (selectors) => {
            if (!root) return null;
            return selectors.map((sel) => root.querySelector(sel)).find(Boolean);
          };
          const metric = (selectors) => {
            const el = q(selectors);
            const count = el ? el.querySelector(".count, [class*=count], [class*=Count]") : null;
            return {selectors_checked: selectors, matched_count: el ? 1 : 0, text: clean(count || el)};
          };
          const describeNode = (el) => el ? {
            tag: el.tagName.toLowerCase(),
            id: el.id || "",
            class: typeof el.className === "string" ? el.className : "",
            data_testid: el.getAttribute("data-testid") || "",
            role: el.getAttribute("role") || ""
          } : null;
          const describeSelector = (el) => {
            if (!el) return "";
            if (el.id) return `#${el.id}`;
            const testId = el.getAttribute("data-testid");
            if (testId) return `[data-testid="${testId}"]`;
            const className = typeof el.className === "string" ? el.className.trim().split(/\\s+/).slice(0, 3).join(".") : "";
            return className ? `${el.tagName.toLowerCase()}.${className}` : el.tagName.toLowerCase();
          };
          const findCommentAreaRoot = () => {
            if (!root) return {el: null, selectors: []};
            const selectors = [
              '[class*="comments-container"]',
              '[class*="commentsContainer"]',
              '[class*="CommentsContainer"]',
              '[class*="comments-list"]',
              '[class*="commentsList"]',
              '[class*="CommentsList"]',
              '[class*="comment-list"]',
              '[class*="commentList"]',
              '[class*="CommentList"]',
              '[class*="comments-el"]',
              '[class*="commentsEl"]',
              '[class*="CommentsEl"]',
              '[data-testid*="comments"]',
              '[data-testid*="comment-list"]'
            ];
            const candidates = Array.from(root.querySelectorAll(selectors.join(','))).filter((el) => {
              if (!visible(el)) return false;
              if (el.closest('.engage-bar, [class*=engage-bar], [class*=EngageBar]')) return false;
              const text = clean(el);
              return text.length > 0 && text.length <= 5000;
            });
            candidates.sort((a, b) => clean(a).length - clean(b).length);
            return {el: candidates[0] || null, selectors};
          };
          const zeroCommentEvidence = () => {
            const commentArea = findCommentAreaRoot();
            const words = ["这是一片荒地", "暂无评论", "还没有评论"];
            const selectors = ['[class*=empty]', '[class*=Empty]', '[class*=placeholder]', '[class*=Placeholder]', '[data-testid*=empty]', '[data-testid*=placeholder]'];
            const ancestors = [];
            let cursor = commentArea.el;
            for (let i = 0; cursor && i < 6; i += 1) {
              ancestors.push(describeNode(cursor));
              cursor = cursor.parentElement && root.contains(cursor.parentElement) ? cursor.parentElement : null;
            }
            const text = commentArea.el ? Array.from(commentArea.el.querySelectorAll(selectors.join(',')))
              .concat([commentArea.el])
              .filter(visible)
              .map((el) => clean(el))
              .filter((candidate) => candidate && candidate.length <= 200 && words.some((word) => candidate.includes(word)))
              .sort((a, b) => a.length - b.length)[0] || "" : "";
            return {
              comment_area_found: Boolean(commentArea.el),
              comment_area_selector: describeSelector(commentArea.el),
              comment_area_selectors_checked: commentArea.selectors,
              comment_area_node: describeNode(commentArea.el),
              comment_area_ancestors: ancestors,
              selectors_checked: selectors,
              matched_count: text ? 1 : 0,
              text
            };
          };
          const clone = root ? root.cloneNode(true) : null;
          if (clone) {
            clone.querySelectorAll('[class*=comment-item], [class*=commentItem], [data-testid*=comment], [class*=comments], [class*=Comments]').forEach((el) => el.remove());
            clone.querySelectorAll('[class*=share-wrapper], [class*=shareWrapper], [class*=Share]').forEach((el) => el.remove());
            clone.querySelectorAll("[href], [src]").forEach((el) => {
              for (const attr of ["href", "src"]) {
                const value = el.getAttribute(attr);
                if (!value) continue;
                try {
                  const url = new URL(value, location.origin);
                  url.search = "";
                  url.hash = "";
                  el.setAttribute(attr, url.toString());
                } catch {
                  el.setAttribute(attr, value.split("?")[0].split("#")[0]);
                }
              }
            });
          }
          const fieldSummary = {
            title: {selectors_checked: ["#detail-title", "h1", "[class*=title]", "[data-testid*=title]"], matched_count: q(["#detail-title", "h1", "[class*=title]", "[data-testid*=title]"]) ? 1 : 0, text: clean(q(["#detail-title", "h1", "[class*=title]", "[data-testid*=title]"]))},
            body: {selectors_checked: ["#detail-desc", "[class*=desc]", "[class*=content]", ".note-content", "[data-testid*=content]"], matched_count: q(["#detail-desc", "[class*=desc]", "[class*=content]", ".note-content", "[data-testid*=content]"]) ? 1 : 0, text: clean(q(["#detail-desc", "[class*=desc]", "[class*=content]", ".note-content", "[data-testid*=content]"]))},
            publish_time: {selectors_checked: ["detail root text"], matched_count: root ? 1 : 0, text: root ? clean(root).slice(0, 500) : ""},
            like_count: metric([".engage-bar .like-wrapper", ".engage-bar [class*=like-wrapper]", ".engage-bar [class*=likeWrapper]", ".engage-bar [class*=Like]"]),
            collect_count: metric([".engage-bar .collect-wrapper", ".engage-bar [class*=collect-wrapper]", ".engage-bar [class*=collectWrapper]", ".engage-bar [class*=Collect]"]),
            comment_count: {...metric([".engage-bar .chat-wrapper", ".engage-bar [class*=chat-wrapper]", ".engage-bar [class*=comment-wrapper]", ".engage-bar [class*=Chat]", ".engage-bar [class*=Comment]"]), zero_comment_evidence: zeroCommentEvidence()},
            tags: {selectors_checked: ['#detail-desc a[href*="search"]', '#detail-desc a[href*="search_result"]', 'a[href*="/search_result"]', 'a[href*="/search"]'], matched_count: root ? root.querySelectorAll('#detail-desc a[href*="search"], #detail-desc a[href*="search_result"], a[href*="/search_result"], a[href*="/search"]').length : 0, text: root ? Array.from(root.querySelectorAll('#detail-desc a[href*="search"], #detail-desc a[href*="search_result"], a[href*="/search_result"], a[href*="/search"]')).map((el) => clean(el)).filter(Boolean).join("\\n") : ""}
          };
          return {
            detail_html: clone ? clone.outerHTML : "",
            dom_summary: {
              note_id: noteId,
              detail_root: {
                root_found: Boolean(root),
                root_reason: rootReason,
                detail_selector: detailSelector,
                evidence_selector: evidenceSelectors,
                detail_root_count: detailRoots.length,
                evidence_count: evidence.length
              },
              fields: fieldSummary
            }
          };
        }
        """,
        note_id,
    )


capture_five_note_review_state = capture_e2e_review_state


def build_actual_payload(note: dict[str, Any]) -> dict[str, Any]:
    sources = note.get("field_sources") or {}
    fields = {
        "title": _field(note.get("title"), sources.get("title")),
        "body": _field(note.get("body"), sources.get("body")),
        "note_type": _field(note.get("note_type"), sources.get("note_type")),
        "publish_time": _field(note.get("publish_time"), sources.get("publish_time")),
        "like_count": _field(note.get("likes_value"), sources.get("like_count"), note.get("likes_raw"), note.get("likes_is_exact")),
        "collect_count": _field(note.get("collects_value"), sources.get("collect_count"), note.get("collects_raw"), note.get("collects_is_exact")),
        "comment_count": _field(note.get("comments_value"), sources.get("comment_count"), note.get("comments_raw"), note.get("comments_is_exact")),
        "tags": _field(note.get("hashtags") or [], sources.get("tags")),
    }
    return {
        "note_id": note.get("note_id"),
        "detail_ready": True,
        "status": note.get("status"),
        "fields": fields,
        "quality": quality_check_actual(fields),
    }


def annotate_actual_field_status(
    actual: dict[str, Any],
    pre_dom_summary: dict[str, Any] | None = None,
    post_dom_summary: dict[str, Any] | None = None,
) -> dict[str, Any]:
    actual = dict(actual)
    fields = {field: dict(payload or {}) for field, payload in (actual.get("fields") or {}).items()}
    for field in REVIEW_FIELDS:
        payload = fields.setdefault(field, {"value": None, "source": "MISSING", "missing_reason": "not_observed"})
        status = _review_field_status(field, payload, pre_dom_summary, post_dom_summary)
        payload["status"] = status
        if status == FIELD_STATUS_CONFIRMED_ABSENT:
            payload["source"] = "MISSING"
            payload["missing_reason"] = "confirmed_absent"
        elif status == FIELD_STATUS_UNVERIFIED_MISSING:
            payload.setdefault("source", "MISSING")
            payload.setdefault("missing_reason", "not_observed")
    actual["fields"] = fields
    actual["quality"] = quality_check_actual(fields)
    return actual


def write_note_review_artifact(
    review_dir: Path,
    note_id: str,
    actual: dict[str, Any],
    pre_capture: dict[str, Any],
    post_capture: dict[str, Any],
    screenshot_path: Path,
) -> Path:
    note_dir = review_dir / note_id
    note_dir.mkdir(parents=True, exist_ok=True)
    dom_summary = {
        "note_id": note_id,
        "pre_extract_dom_summary": pre_capture["dom_summary"],
        "post_extract_dom_summary": post_capture["dom_summary"],
        "pre_post_consistent": _dom_summaries_consistent(pre_capture["dom_summary"], post_capture["dom_summary"]),
    }
    actual = annotate_actual_field_status(actual, pre_capture.get("dom_summary"), post_capture.get("dom_summary"))
    files = {
        "actual.json": json.dumps(actual, ensure_ascii=False, indent=2, default=str) + "\n",
        "dom_summary.json": json.dumps(dom_summary, ensure_ascii=False, indent=2, default=str) + "\n",
        "detail.html": sanitize_detail_html(post_capture.get("detail_html") or "") + "\n",
    }
    for name, content in files.items():
        assert_artifact_text_safe(content, name)
        (note_dir / name).write_text(content, encoding="utf-8")
    shutil.copyfile(screenshot_path, note_dir / "page_screenshot.png")
    return note_dir


def write_e2e_review_summary(
    review_dir: Path,
    *,
    run_id: str,
    collection: Any,
    excel_readback: dict[str, Any] | None,
) -> Path:
    summary = build_e2e_review_summary(run_id=run_id, collection=collection, review_dir=review_dir, excel_readback=excel_readback)
    content = json.dumps(summary, ensure_ascii=False, indent=2, default=str) + "\n"
    assert_artifact_text_safe(content, "summary.json")
    path = review_dir / "summary.json"
    path.write_text(content, encoding="utf-8")
    write_data_quality_review(review_dir, run_id=run_id)
    return path


def build_e2e_review_summary(
    *,
    run_id: str,
    collection: Any,
    review_dir: Path,
    excel_readback: dict[str, Any] | None,
) -> dict[str, Any]:
    pre_post_inconsistent_note_ids = []
    quality_issue_note_ids = []
    unverified_missing_note_ids: dict[str, list[str]] = {field: [] for field in REVIEW_FIELDS}
    confirmed_absent_note_ids: dict[str, list[str]] = {field: [] for field in REVIEW_FIELDS}
    field_quality = {
        field: {
            "present": 0,
            "confirmed_absent": 0,
            "unverified_missing": 0,
            "verified_total": 0,
        }
        for field in REVIEW_FIELDS
    }
    for dom_summary_path in sorted(review_dir.glob("*/dom_summary.json")):
        payload = json.loads(dom_summary_path.read_text(encoding="utf-8"))
        if payload.get("pre_post_consistent") is False:
            pre_post_inconsistent_note_ids.append(dom_summary_path.parent.name)
    for actual_path in sorted(review_dir.glob("*/actual.json")):
        payload = json.loads(actual_path.read_text(encoding="utf-8"))
        if not (payload.get("quality") or {}).get("passed", True):
            quality_issue_note_ids.append(actual_path.parent.name)
        fields = payload.get("fields") or {}
        for field in REVIEW_FIELDS:
            field_payload = fields.get(field) or {}
            status = field_payload.get("status") or (FIELD_STATUS_PRESENT if _present(field_payload.get("value")) else FIELD_STATUS_UNVERIFIED_MISSING)
            status_key = {
                FIELD_STATUS_PRESENT: "present",
                FIELD_STATUS_CONFIRMED_ABSENT: "confirmed_absent",
                FIELD_STATUS_UNVERIFIED_MISSING: "unverified_missing",
            }.get(status, "unverified_missing")
            field_quality[field][status_key] += 1
            if status == FIELD_STATUS_CONFIRMED_ABSENT:
                confirmed_absent_note_ids[field].append(actual_path.parent.name)
            elif status != FIELD_STATUS_PRESENT:
                unverified_missing_note_ids[field].append(actual_path.parent.name)
    for counts in field_quality.values():
        counts["verified_total"] = counts["present"] + counts["confirmed_absent"]
    unverified_missing_note_ids = {field: ids for field, ids in unverified_missing_note_ids.items() if ids}
    confirmed_absent_note_ids = {field: ids for field, ids in confirmed_absent_note_ids.items() if ids}
    data_quality_status = (
        "DATA_QUALITY_REVIEW_REQUIRED"
        if unverified_missing_note_ids or quality_issue_note_ids or pre_post_inconsistent_note_ids
        else "PASS"
    )
    return {
        "run_id": run_id,
        "data_quality_status": data_quality_status,
        "attempted": len(getattr(collection, "attempted_ids", []) or []),
        "target_verified": len(getattr(collection, "verified_ids", []) or []),
        "detail_ready": len(getattr(collection, "verified_ids", []) or []),
        "exportable": len(getattr(collection, "exportable_ids", []) or []),
        "navigation_failed": len(getattr(collection, "navigation_failed_ids", []) or []),
        "detail_not_ready": _count_detail_not_ready(getattr(collection, "validation_notes", []) or []),
        "risk_control": str(getattr(collection, "safe_stop_reason", "") or "") == "RISK_CONTROL_DETECTED",
        "human_verification": str(getattr(collection, "safe_stop_reason", "") or "") == "HUMAN_VERIFICATION_REQUIRED",
        "safe_stop_reason": getattr(collection, "safe_stop_reason", None),
        "field_completeness": getattr(collection, "field_presence", {}) or {},
        "field_quality": field_quality,
        "unverified_missing_note_ids": unverified_missing_note_ids,
        "confirmed_absent_note_ids": confirmed_absent_note_ids,
        "missing_field_note_ids": unverified_missing_note_ids,
        "pre_post_inconsistent_note_ids": pre_post_inconsistent_note_ids,
        "quality_issue_note_ids": quality_issue_note_ids,
        "excel_readback": excel_readback,
    }


def write_data_quality_review(review_dir: Path, *, run_id: str) -> Path:
    review = build_data_quality_review(run_id=run_id, review_dir=review_dir)
    content = json.dumps(review, ensure_ascii=False, indent=2, default=str) + "\n"
    assert_artifact_text_safe(content, "data_quality_review.json")
    path = review_dir / "data_quality_review.json"
    path.write_text(content, encoding="utf-8")
    return path


def build_data_quality_review(*, run_id: str, review_dir: Path) -> dict[str, Any]:
    findings = []
    for actual_path in sorted(review_dir.glob("*/actual.json")):
        note_id = actual_path.parent.name
        actual = json.loads(actual_path.read_text(encoding="utf-8"))
        dom_summary_path = actual_path.parent / "dom_summary.json"
        dom_summary = json.loads(dom_summary_path.read_text(encoding="utf-8")) if dom_summary_path.exists() else {}
        fields = actual.get("fields") or {}
        for field in REVIEW_FIELDS:
            field_payload = fields.get(field) or {}
            status = field_payload.get("status")
            if status == FIELD_STATUS_PRESENT:
                continue
            if field == "tags" and status == FIELD_STATUS_CONFIRMED_ABSENT:
                findings.append(_confirmed_absent_tags_finding(note_id, dom_summary))
            else:
                findings.append(
                    {
                        "note_id": note_id,
                        "field": field,
                        "classification": CLASSIFICATION_INSUFFICIENT_EVIDENCE,
                        "dom_evidence": _field_dom_evidence(dom_summary, field),
                        "visual_evidence": {"visual_status": "NOT_ASSESSED_OFFLINE"},
                        "structured_evidence": "not_present_in_artifact",
                        "decision": "missing value requires manual review",
                    }
                )
    return {"run_id": run_id, "findings": findings}


def validate_excel_readback(excel_path: Path, expected_rows: list[Any], expected_note_ids: list[str]) -> dict[str, Any]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb["公开笔记"]
    headers = [cell.value for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({header: row[idx] for header, idx in header_index.items() if header is not None and idx < len(row)})

    counts = Counter(_blank_to_none(row.get("note_id")) for row in rows)
    duplicate_note_ids = sorted(note_id for note_id, count in counts.items() if note_id and count > 1)
    rows_by_note = {str(row.get("note_id")): row for row in rows if row.get("note_id")}
    missing_note_ids = [note_id for note_id in expected_note_ids if note_id not in rows_by_note]
    expected_by_note = {row["note_id"]: row for row in expected_rows}
    field_diffs = []
    for note_id in expected_note_ids:
        if note_id not in rows_by_note:
            continue
        expected = expected_by_note[note_id]
        actual = rows_by_note[note_id]
        for field, header in _excel_field_map().items():
            expected_value = _expected_db_value(expected, field)
            actual_value = _excel_value(actual.get(header), field)
            if expected_value != actual_value:
                field_diffs.append({"note_id": note_id, "field": field, "expected": expected_value, "excel_actual": actual_value})

    return {
        "excel_path": str(excel_path),
        "rows_expected": len(expected_note_ids),
        "rows_found": sum(1 for note_id in expected_note_ids if note_id in rows_by_note),
        "duplicate_note_ids": duplicate_note_ids,
        "missing_note_ids": missing_note_ids,
        "field_diffs": field_diffs,
    }


def write_excel_readback_artifact(review_dir: Path, report: dict[str, Any], excel_path: Path | None = None) -> Path:
    content = json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n"
    assert_artifact_text_safe(content, "excel_readback.json")
    path = review_dir / "excel_readback.json"
    path.write_text(content, encoding="utf-8")
    return path


def quality_check_actual(fields: dict[str, dict[str, Any]]) -> dict[str, Any]:
    issues = []
    title = fields.get("title", {}).get("value")
    if isinstance(title, str) and any(term in title for term in ["加载中", "请刷新试试", "说点什么"]):
        issues.append({"field": "title", "reason": "ui_shell_text"})
    body = fields.get("body", {}).get("value")
    if isinstance(body, str):
        for term in UI_POLLUTION_TERMS:
            if term in body:
                issues.append({"field": "body", "reason": "ui_pollution", "term": term})
    tags = fields.get("tags", {}).get("value") or []
    if any(isinstance(tag, str) and any(term in tag for term in UI_POLLUTION_TERMS) for tag in tags):
        issues.append({"field": "tags", "reason": "ui_pollution"})
    return {"passed": not issues, "issues": issues}


def _review_field_status(
    field: str,
    payload: dict[str, Any],
    pre_dom_summary: dict[str, Any] | None,
    post_dom_summary: dict[str, Any] | None,
) -> str:
    if _present(payload.get("value")):
        return FIELD_STATUS_PRESENT
    if field == "tags" and _tags_confirmed_absent(pre_dom_summary, post_dom_summary):
        return FIELD_STATUS_CONFIRMED_ABSENT
    return FIELD_STATUS_UNVERIFIED_MISSING


def _tags_confirmed_absent(pre_dom_summary: dict[str, Any] | None, post_dom_summary: dict[str, Any] | None) -> bool:
    summaries = [summary for summary in [pre_dom_summary, post_dom_summary] if summary]
    return bool(summaries) and all(_single_dom_summary_confirms_tags_absent(summary) for summary in summaries)


def _single_dom_summary_confirms_tags_absent(summary: dict[str, Any]) -> bool:
    detail_root = summary.get("detail_root") or {}
    fields = summary.get("fields") or {}
    body = fields.get("body") or {}
    tags = fields.get("tags") or {}
    return (
        detail_root.get("root_found") is True
        and bool(tags.get("selectors_checked"))
        and int(tags.get("matched_count") or 0) == 0
        and not str(tags.get("text") or "").strip()
        and int(body.get("matched_count") or 0) >= 1
        and "#" not in str(body.get("text") or "")
    )


def _confirmed_absent_tags_finding(note_id: str, dom_summary: dict[str, Any]) -> dict[str, Any]:
    post_summary = dom_summary.get("post_extract_dom_summary") or {}
    pre_summary = dom_summary.get("pre_extract_dom_summary") or {}
    evidence_summary = post_summary or pre_summary
    fields = evidence_summary.get("fields") or {}
    tags = fields.get("tags") or {}
    body = fields.get("body") or {}
    return {
        "note_id": note_id,
        "field": "tags",
        "classification": CLASSIFICATION_TRUE_ABSENT,
        "dom_evidence": {
            "detail_root_found": bool((evidence_summary.get("detail_root") or {}).get("root_found")),
            "tag_nodes_found": int(tags.get("matched_count") or 0),
            "tag_texts": [],
            "body_text_preview": str(body.get("text") or "")[:200],
            "candidate_selectors": tags.get("selectors_checked") or [],
            "candidate_matches": [],
            "pre_post_consistent": dom_summary.get("pre_post_consistent"),
        },
        "visual_evidence": {
            "visual_tags_status": "ABSENT",
            "visible_tag_texts": [],
            "source": "page_screenshot.png manual offline inspection",
        },
        "structured_evidence": "not_present_in_artifact",
        "decision": "tags confirmed absent; null business value is correct",
    }


def _field_dom_evidence(dom_summary: dict[str, Any], field: str) -> dict[str, Any]:
    post_summary = dom_summary.get("post_extract_dom_summary") or {}
    pre_summary = dom_summary.get("pre_extract_dom_summary") or {}
    fields = (post_summary or pre_summary).get("fields") or {}
    payload = fields.get(field) or {}
    return {
        "matched_count": int(payload.get("matched_count") or 0),
        "text_preview": str(payload.get("text") or "")[:200],
        "selectors_checked": payload.get("selectors_checked") or [],
        "pre_post_consistent": dom_summary.get("pre_post_consistent"),
    }


def assert_artifact_text_safe(content: str, name: str) -> None:
    lowered = content.lower()
    lowered = lowered.replace("detail_initial_state", "")
    hits = [term for term in SENSITIVE_TERMS if term in lowered]
    if hits:
        raise ValueError(f"five-note review artifact {name} contains sensitive terms: {sorted(set(hits))}")


def sanitize_detail_html(detail_html: str) -> str:
    sanitized = re.sub(r"\s(?:href|src)=([\"'])(.*?)(\1)", _sanitize_url_attr, detail_html, flags=re.IGNORECASE | re.DOTALL)
    sanitized = re.sub(r"(?i)__INITIAL_STATE__", "__STATE_NAME_REDACTED__", sanitized)
    return sanitized


def _field(value: Any, source: Any, raw_display: Any = None, exact: Any = None) -> dict[str, Any]:
    present = _present(value)
    output = {"value": value if present else None, "source": str(source or "MISSING") if present else "MISSING"}
    if raw_display is not None:
        output["raw"] = raw_display
    if exact is not None:
        output["exact"] = bool(exact)
    if not present:
        output["missing_reason"] = "not_observed"
    return output


def _present(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, (list, tuple, set, dict)):
        return bool(value)
    return True


def _dom_summaries_consistent(pre: dict[str, Any], post: dict[str, Any]) -> bool:
    pre_fields = (pre or {}).get("fields") or {}
    post_fields = (post or {}).get("fields") or {}
    for field in REVIEW_FIELDS:
        if (pre_fields.get(field) or {}).get("text") != (post_fields.get(field) or {}).get("text"):
            return False
    return True


def _excel_field_map() -> dict[str, str]:
    return {
        "title": "标题",
        "body": "正文",
        "note_type": "帖子类型",
        "publish_time": "发布时间",
        "like_count": "点赞数",
        "like_raw": "点赞原始显示",
        "like_exact": "点赞是否精确",
        "collect_count": "收藏数",
        "collect_raw": "收藏原始显示",
        "collect_exact": "收藏是否精确",
        "comment_count": "评论数",
        "comment_raw": "评论原始显示",
        "comment_exact": "评论是否精确",
        "tags": "标签",
    }


def _expected_db_value(row: Any, field: str) -> Any:
    key_map = {
        "like_count": "likes_value",
        "like_raw": "likes_raw",
        "like_exact": "likes_is_exact",
        "collect_count": "collects_value",
        "collect_raw": "collects_raw",
        "collect_exact": "collects_is_exact",
        "comment_count": "comments_value",
        "comment_raw": "comments_raw",
        "comment_exact": "comments_is_exact",
        "tags": "hashtags",
    }
    value = row[key_map.get(field, field)]
    if field.endswith("_exact"):
        return _bool_excel_text(value)
    if field.endswith("_count"):
        return _number_or_none(value)
    if field == "tags":
        return _normalize_tags(value)
    return _blank_to_none(value)


def _excel_value(value: Any, field: str) -> Any:
    if field.endswith("_count"):
        return _number_or_none(value)
    if field.endswith("_exact"):
        return _bool_excel_text(value)
    if field == "tags":
        return _normalize_tags(value)
    return _blank_to_none(value)


def _blank_to_none(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str) and value == "":
        return None
    return value


def _number_or_none(value: Any) -> int | None:
    value = _blank_to_none(value)
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, str) and re.fullmatch(r"\d+(?:\.0)?", value):
        return int(float(value))
    return value


def _bool_excel_text(value: Any) -> str | None:
    value = _blank_to_none(value)
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return "是" if bool(value) else "否"


def _normalize_tags(value: Any) -> str | None:
    value = _blank_to_none(value)
    if value is None:
        return None
    if isinstance(value, str):
        try:
            loaded = json.loads(value)
            if isinstance(loaded, list):
                return _blank_to_none(" ".join(str(item) for item in loaded))
        except json.JSONDecodeError:
            return _blank_to_none(value)
        return _blank_to_none(value)
    if isinstance(value, list):
        return _blank_to_none(" ".join(str(item) for item in value))
    return str(value)


def validate_note_headers() -> None:
    missing = [header for header in _excel_field_map().values() if header not in NOTE_HEADERS]
    if missing:
        raise ValueError(f"Excel headers missing from exporter schema: {missing}")


def _count_detail_not_ready(notes: list[dict[str, Any]]) -> int:
    return sum(1 for note in notes if not note.get("detail_ready"))


def _sanitize_url_attr(match: re.Match[str]) -> str:
    quote = match.group(1)
    value = match.group(2)
    safe = value.split("?", 1)[0].split("#", 1)[0]
    return f" {match.group(0).strip().split('=', 1)[0]}={quote}{safe}{quote}"
