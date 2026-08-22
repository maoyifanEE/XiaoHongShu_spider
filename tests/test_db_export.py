from pathlib import Path

from openpyxl import load_workbook

from xhs_profile_exporter.db import Database
from xhs_profile_exporter.exporter import NOTE_HEADERS, export_excel
from xhs_profile_exporter.qa import run_offline_qa


class DummyLogger:
    def info(self, *args, **kwargs):
        pass


def test_db_qa_and_export(tmp_path: Path):
    db = Database(tmp_path / "data" / "xhs_data.sqlite3")
    db.migrate()
    creator_id = "creator1"
    db.save_profile_snapshot(
        creator_id,
        {
            "captured_at": "2026-08-18T20:00:00+08:00",
            "nickname": "测试博主",
            "xhs_id": "test",
            "canonical_url": "https://www.xiaohongshu.com/user/profile/creator1",
            "followers_value": 100,
            "followers_raw": "100",
            "followers_is_exact": True,
            "source": "test",
        },
    )
    db.upsert_note(
        creator_id,
        {
            "note_id": "note1",
            "canonical_url": "https://www.xiaohongshu.com/explore/note1",
            "title": "标题",
            "note_type": "图文",
            "publish_time": "2026-08-18",
            "body": "正文 #标签",
            "hashtags": ["#标签"],
            "likes_value": 1,
            "likes_raw": "1",
            "likes_is_exact": True,
            "collects_value": None,
            "collects_raw": None,
            "collects_is_exact": None,
            "comments_value": 1,
            "comments_raw": "1",
            "comments_is_exact": True,
            "top_comments": [{"rank": 1, "author_name": "A", "body": "评论", "source": "test"}],
            "status": "OK",
            "source": "test",
        },
    )
    report = run_offline_qa(db, creator_id, DummyLogger())
    assert report["passed"] is True
    assert "shares" not in report["metrics"]
    assert "comments" not in report
    assert db.conn.execute("SELECT COUNT(*) FROM top_comments").fetchone()[0] == 0
    metric_columns = {
        row[1]
        for row in db.conn.execute("PRAGMA table_info(note_metrics_snapshots)").fetchall()
    }
    assert {"shares_value", "shares_raw", "shares_is_exact"}.issubset(metric_columns)
    metrics = db.conn.execute(
        "SELECT shares_value, shares_raw, shares_is_exact FROM note_metrics_snapshots WHERE note_id = ?",
        ("note1",),
    ).fetchone()
    assert tuple(metrics) == (None, None, None)
    current = db.current_notes(creator_id)[0]
    assert "shares_value" not in current.keys()
    path = export_excel(db, tmp_path, creator_id, "测试博主", DummyLogger())
    wb = load_workbook(path)
    assert "博主主页" in wb.sheetnames
    assert "公开笔记" in wb.sheetnames
    assert wb["公开笔记"].max_row == 2
    headers = [cell.value for cell in wb["公开笔记"][1]]
    assert headers == NOTE_HEADERS
    assert "分享数" not in headers
    assert "分享原始显示" not in headers
    assert "分享是否精确" not in headers
    assert not any(str(header).startswith(("评论1_", "评论2_", "评论3_")) for header in headers)
    assert {"评论数", "评论原始显示", "评论是否精确"}.issubset(set(headers))
    db.close()
